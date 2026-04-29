import openpyxl
from openpyxl.styles import Font, Alignment
from io import BytesIO
from datetime import datetime

def create_excel_receipt(user, items, total_price, phone=None, address=None, order_id=None):
    """Создание Excel чека"""
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Чек заказа"
    
    # Заголовок
    ws['A1'] = 'ЧЕК ЗАКАЗА'
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A1:D1')
    
    # Номер заказа
    if order_id:
        ws['A3'] = f'Номер заказа: #{order_id}'
    else:
        ws['A3'] = f'Номер заказа: {datetime.now().strftime("%Y%m%d%H%M%S")}'
    
    ws['A4'] = f'Дата: {datetime.now().strftime("%d.%m.%Y %H:%M")}'
    ws['A5'] = f'Покупатель: {user.get_full_name() or user.username}'
    ws['A6'] = f'Email: {user.email}'
    
    row = 7
    if phone:
        ws[f'A{row}'] = f'Телефон: {phone}'
        row += 1
    if address:
        ws[f'A{row}'] = f'Адрес: {address}'
        row += 1
    
    # Пустая строка перед таблицей
    row += 1
    ws[f'A{row}'] = ''
    
    # Заголовки таблицы
    row += 1
    ws[f'A{row}'] = '№'
    ws[f'B{row}'] = 'Товар'
    ws[f'C{row}'] = 'Цена'
    ws[f'D{row}'] = 'Кол-во'
    ws[f'E{row}'] = 'Сумма'
    
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws[f'{col}{row}'].font = Font(bold=True)
        ws[f'{col}{row}'].alignment = Alignment(horizontal='center')
    
    # Заполняем товары
    data_row = row + 1
    for idx, item in enumerate(items, 1):
        ws[f'A{data_row}'] = idx
        ws[f'B{data_row}'] = item.product.name
        ws[f'C{data_row}'] = f'{item.product.price:,.2f} ₽'
        ws[f'D{data_row}'] = item.count
        ws[f'E{data_row}'] = f'{item.product.price * item.count:,.2f} ₽'
        
        ws[f'A{data_row}'].alignment = Alignment(horizontal='center')
        ws[f'C{data_row}'].alignment = Alignment(horizontal='right')
        ws[f'D{data_row}'].alignment = Alignment(horizontal='center')
        ws[f'E{data_row}'].alignment = Alignment(horizontal='right')
        data_row += 1
    
    # Итог
    data_row += 1
    ws[f'D{data_row}'] = 'ИТОГО:'
    ws[f'D{data_row}'].font = Font(bold=True)
    ws[f'D{data_row}'].alignment = Alignment(horizontal='right')
    
    ws[f'E{data_row}'] = f'{total_price:,.2f} ₽'
    ws[f'E{data_row}'].font = Font(bold=True)
    ws[f'E{data_row}'].alignment = Alignment(horizontal='right')
    
    # Нижний колонтитул
    data_row += 2
    ws[f'A{data_row}'] = 'Спасибо за покупку!'
    ws[f'A{data_row}'].alignment = Alignment(horizontal='center')
    ws.merge_cells(f'A{data_row}:E{data_row}')
    
    # Настраиваем ширину колонок
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 15
    
    # Сохраняем
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output